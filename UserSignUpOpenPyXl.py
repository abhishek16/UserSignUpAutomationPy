#for x in range(5):


import time, sys, time, os, pyautogui, webbrowser, openpyxl, pandas as pd
from ddt import ddt, data, unpack
from openpyxl.cell.read_only import EmptyCell
from selenium import webdriver


excel_document = openpyxl.load_workbook(r"C:\Users\akulkarni\Desktop\Arroweye_Customer Service Tool Test.xlsx", data_only=True)    
sheet = excel_document["TestDirectory"]
print (sheet.max_row)
for i in range(1, sheet.max_row+1):

        driver = webdriver.Chrome(executable_path="C:/chromedriver_win32/chromedriver.exe")

        # URL declaration
        base_url="https://portalwebt1.asitest.net/Portal3/Login?ReturnUrl=%2fPortal3%2fUserManagement%2fUserEdit%3fClientId%3d152&ClientId=152"
        driver.maximize_window()
        driver.get(base_url)

        #Login action
        LoginTextBox = driver.find_element_by_id("UserName")
        LoginTextBox.clear()
        LoginTextBox.send_keys("")
        PasswordTextBox = driver.find_element_by_id("Password")
        PasswordTextBox.clear()
        PasswordTextBox.send_keys("")
        LoginButton = driver.find_element_by_xpath("/html/body/div[2]/div/div/div/div/form/fieldset/div[4]/div/button")
        LoginButton.click()

        time.sleep(2)

        #Form fillup begins here

        cell = "A" + str(i)
        userName = sheet[cell].value
        AddName = driver.find_element_by_id("txtUserName")
        AddName.send_keys(userName)


        cell = "B" + str(i)
        emailAddress = sheet[cell].value 
        AddEmail = driver.find_element_by_id("Email")
        AddEmail.send_keys(emailAddress)

        time.sleep(2)

        #Active status
        MakeActive = driver.find_element_by_xpath("/html/body/div[1]/div[2]/form/div/fieldset[1]/div[4]/label").click()


        #Name & Address
        cell = "C" + str(i)
        userFName = sheet[cell].value   
        AddFName = driver.find_element_by_id("FirstName")
        AddFName.send_keys(userFName)

        cell = "D" + str(i)
        userLName = sheet[cell].value  
        AddLName = driver.find_element_by_xpath("/html/body/div[1]/div[2]/form/div/fieldset[2]/div[3]/div/input")
        AddLName.send_keys(userLName)

        
        AddLine1 = driver.find_element_by_id("Address1")
        AddLine1.send_keys("400 W. Covina Blvd")
        AddCity = driver.find_element_by_id("City")
        AddCity.send_keys("San Dimas")
        AddState = driver.find_element_by_id("State")
        AddState.send_keys("CA")
        AddCountry = driver.find_element_by_id("Country")
        AddCountry.send_keys("USA")
        AddZipCode = driver.find_element_by_id("ZipCode")
        AddZipCode.send_keys("91733")


        #Select user role here
        driver.find_element_by_xpath("//select[@name='UsersSecurityRoleForCurrentClient']/option[text()='CLIENT_SERVICES_PORTAL_USER']").click()

        #Set Password & Confirm Password here
        SetPassword = driver.find_element_by_id("Password")
        SetPassword.send_keys("asdfgfqwerr432321@@!")

        ConfPassword = driver.find_element_by_id("ConfirmPassword")
        ConfPassword.send_keys("asdfgfqwerr432321@@!")


        time.sleep(1)
        
        #Submit user form
        driver.find_element_by_xpath("//*[@id='divContent']/form/div/div/div/input[1]").click()
        time.sleep(1)

        driver.find_element_by_xpath("/html/body/nav/div/div[2]/ul[2]/li[2]/a/span").click()
        time.sleep(1)

        driver.close()


#User creation count
        print ("User {} - {} created successfully!".format(i, userName))
else:
        print("All users created successfully!")

